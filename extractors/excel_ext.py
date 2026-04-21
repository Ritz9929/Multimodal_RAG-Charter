"""
Excel Extractor
================
Uses openpyxl (via pandas) for Excel parsing with multi-sheet support
and intelligent row-windowing.

Strategy:
  Page 0: Workbook-level summary (sheet names, row/col counts, overview)
  Per Sheet:
    - Sheet schema page (column names, types, stats)
    - Row windows as markdown tables (ROWS_PER_PAGE per page)
  Embedded Images: Extracted from each sheet via openpyxl

Why multi-sheet awareness?
  Excel files often have data spread across multiple sheets (e.g.,
  "Revenue", "Expenses", "Summary"). Treating the file as a flat table
  would lose this context. Sheet names are preserved as section headers
  so the RAG system can answer queries like "what's on the Revenue sheet?".
"""

import logging
from pathlib import Path
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

from .base import BaseExtractor, ExtractionResult, ExtractedImage, ExtractedTable

logger = logging.getLogger(__name__)


class ExcelExtractor(BaseExtractor):
    """
    Excel (.xlsx / .xls) extraction with multi-sheet support.

    Strategy:
      Page 0: Workbook overview (sheets, total rows, column summaries)
      Per sheet:
        - Schema summary page
        - Row-windowed markdown tables (ROWS_PER_PAGE rows per page)
      Embedded images extracted via openpyxl.
    """

    ROWS_PER_PAGE = 50  # Rows per "page" (matches CSV extractor default)

    def __init__(self, output_dir: str = "./mock_s3_storage",
                 rows_per_page: int = 50, **kwargs):
        super().__init__(output_dir)
        self.ROWS_PER_PAGE = rows_per_page

    # ── Schema summary builders ──────────────────────────────────────────

    def _build_workbook_summary(self, sheets_info: list[dict],
                                file_path: str) -> str:
        """Build workbook-level overview for Page 0."""
        fname = Path(file_path).name
        total_rows = sum(s["rows"] for s in sheets_info)
        total_sheets = len(sheets_info)

        lines = [
            f"# Excel Workbook Summary: {fname}",
            f"",
            f"**Total Sheets**: {total_sheets}",
            f"**Total Rows (all sheets)**: {total_rows}",
            f"",
            f"## Sheet Index",
        ]

        for i, info in enumerate(sheets_info, 1):
            lines.append(
                f"- **Sheet {i}: \"{info['name']}\"** — "
                f"{info['rows']} rows × {info['cols']} columns"
            )
            if info["columns"]:
                col_preview = ", ".join(str(c) for c in info["columns"][:8])
                if len(info["columns"]) > 8:
                    col_preview += f" … (+{len(info['columns']) - 8} more)"
                lines.append(f"  - Columns: {col_preview}")

        return "\n".join(lines)

    def _build_sheet_schema(self, df: pd.DataFrame, sheet_name: str,
                            sheet_idx: int, total_sheets: int) -> str:
        """Build per-sheet schema summary (mirrors CSVExtractor style)."""
        lines = [
            f"# Sheet {sheet_idx}/{total_sheets}: \"{sheet_name}\"",
            f"",
            f"**Rows**: {len(df)}",
            f"**Columns**: {len(df.columns)}",
            f"",
            f"## Column Definitions",
        ]

        for col in df.columns:
            dtype = str(df[col].dtype)
            non_null = df[col].notna().sum()
            null_pct = (
                (df[col].isna().sum() / len(df)) * 100 if len(df) > 0 else 0
            )

            lines.append(
                f"- **{col}** ({dtype}): "
                f"{non_null} non-null values ({null_pct:.1f}% missing)"
            )

            # Sample unique values
            samples = df[col].dropna().unique()[:5]
            if len(samples) > 0:
                sample_str = ", ".join(str(s) for s in samples)
                lines.append(f"  - Sample values: {sample_str}")

        # Numeric statistics
        numeric_cols = df.select_dtypes(include="number").columns
        if len(numeric_cols) > 0:
            lines.append("")
            lines.append("## Numeric Column Statistics")
            stats_df = df[numeric_cols].describe()
            lines.append(stats_df.to_markdown())

        return "\n".join(lines)

    # ── Image extraction ─────────────────────────────────────────────────

    def _extract_images(self, wb, doc_dir: Path,
                        source_doc: str) -> list[ExtractedImage]:
        """Extract embedded images from all sheets via openpyxl."""
        images = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            if not hasattr(ws, "_images") or not ws._images:
                continue

            for img_idx, img in enumerate(ws._images):
                try:
                    # openpyxl stores image data in img._data or img.ref
                    if hasattr(img, "_data") and img._data:
                        img_data = img._data()
                    elif hasattr(img, "ref"):
                        img_data = img.ref
                    else:
                        continue

                    # Determine format from content type or default to png
                    ext = "png"
                    if hasattr(img, "format") and img.format:
                        ext = img.format.lower()

                    safe_sheet = sheet_name.replace(" ", "_").replace("/", "_")
                    filename = f"{safe_sheet}_img{img_idx}.{ext}"
                    filepath = doc_dir / filename

                    if isinstance(img_data, bytes):
                        filepath.write_bytes(img_data)
                    elif hasattr(img_data, "read"):
                        filepath.write_bytes(img_data.read())
                    else:
                        continue

                    sheet_idx = wb.sheetnames.index(sheet_name)
                    images.append(ExtractedImage(
                        filename=filename,
                        filepath=str(filepath),
                        page_number=sheet_idx,
                        position_index=img_idx,
                        source_type="embedded",
                    ))
                    logger.info(f"  Extracted image: {filename} "
                                f"(sheet: {sheet_name})")

                except Exception as e:
                    logger.warning(
                        f"  ⚠ Could not extract image {img_idx} from "
                        f"sheet '{sheet_name}': {e}"
                    )

        return images

    # ── Main extraction ──────────────────────────────────────────────────

    def extract(self, file_path: str, source_doc: str = "") -> ExtractionResult:
        """Parse the Excel file and return schema + windowed rows per sheet."""
        result = ExtractionResult(source_format="excel")

        if not source_doc:
            source_doc = Path(file_path).name

        doc_dir = self._make_doc_output_dir(source_doc)

        # ── Read all sheets with pandas ──
        ext = Path(file_path).suffix.lower()
        engine = "openpyxl" if ext == ".xlsx" else "xlrd"

        try:
            all_sheets: dict[str, pd.DataFrame] = pd.read_excel(
                file_path, sheet_name=None, engine=engine
            )
        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            raise

        sheet_names = list(all_sheets.keys())
        logger.info(
            f"Opened Excel: {file_path} — "
            f"{len(sheet_names)} sheet(s): {sheet_names}"
        )

        # ── Collect sheet metadata for workbook summary ──
        sheets_info = []
        for name, df in all_sheets.items():
            sheets_info.append({
                "name": name,
                "rows": len(df),
                "cols": len(df.columns),
                "columns": list(df.columns),
            })

        # ── Page 0: Workbook overview ──
        result.page_texts[0] = self._build_workbook_summary(
            sheets_info, file_path
        )

        # ── Per-sheet: Schema + row-windowed pages ──
        page_counter = 1  # Next available page number

        for sheet_idx, (sheet_name, df) in enumerate(all_sheets.items(), 1):
            # Skip completely empty sheets
            if df.empty:
                logger.info(f"  Skipping empty sheet: '{sheet_name}'")
                continue

            # Schema page for this sheet
            result.page_texts[page_counter] = self._build_sheet_schema(
                df, sheet_name, sheet_idx, len(sheet_names)
            )
            page_counter += 1

            # Row-windowed markdown tables
            for i, start in enumerate(range(0, len(df), self.ROWS_PER_PAGE)):
                window = df.iloc[start : start + self.ROWS_PER_PAGE]
                md_table = window.to_markdown(index=False)

                header = (
                    f"## Sheet: \"{sheet_name}\" — "
                    f"Rows {start + 1}–{start + len(window)} "
                    f"of {len(df)}\n\n"
                )
                result.page_texts[page_counter] = header + md_table

                result.tables.append(ExtractedTable(
                    markdown=md_table,
                    page_number=page_counter,
                    position_index=0,
                    row_count=len(window),
                    col_count=len(window.columns),
                    caption=f"Sheet: {sheet_name}",
                ))

                page_counter += 1

        # ── Extract embedded images (openpyxl only → .xlsx) ──
        if ext == ".xlsx":
            try:
                wb = load_workbook(file_path, data_only=True)
                result.images = self._extract_images(wb, doc_dir, source_doc)
                wb.close()
            except Exception as e:
                logger.warning(f"  ⚠ Image extraction failed: {e}")

        # ── Metadata ──
        result.metadata = {
            "total_sheets": len(sheet_names),
            "sheet_names": sheet_names,
            "total_rows": sum(s["rows"] for s in sheets_info),
            "total_columns_per_sheet": {
                s["name"]: s["cols"] for s in sheets_info
            },
            "pages_created": len(result.page_texts),
            "rows_per_page": self.ROWS_PER_PAGE,
            "images_extracted": len(result.images),
        }

        logger.info(
            f"Excel extraction complete — "
            f"{len(sheet_names)} sheets, "
            f"{len(result.page_texts)} pages, "
            f"{len(result.images)} images"
        )
        return result
